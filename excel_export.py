"""
excel_export.py — Build a two-sheet XLSX workbook from a list of PO records.

Used by the FastAPI `/api/pos/export.xlsx` endpoint to generate the rolling
Excel ledger from the SQLite database.
"""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)


def _style_header_row(ws, columns: list[tuple[str, int]]):
    for idx, (label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _flatten(s) -> str:
    if not s:
        return ""
    return str(s).replace("\n", " · ").strip()


def build_workbook(records: Iterable[dict]) -> io.BytesIO:
    """Return a BytesIO with a fresh XLSX containing two sheets."""
    records = list(records)

    wb = Workbook()

    # ---------- Sheet 1: Purchase Orders ----------
    ws1 = wb.active
    ws1.title = "Purchase Orders"

    columns_po = [
        ("PO Number", 16),
        ("PO Date", 12),
        ("Revision", 8),
        ("Customer", 28),
        ("Supplier", 28),
        ("Supplier #", 12),
        ("Bill To", 36),
        ("Ship To", 36),
        ("Payment Terms", 16),
        ("Freight Terms", 22),
        ("Ship Via", 14),
        ("F.O.B.", 16),
        ("Buyer", 18),
        ("Buyer Email", 26),
        ("Buyer Phone", 16),
        ("Receiving Contact", 22),
        ("Receiving Phone", 16),
        ("Quote #", 16),
        ("Contract #", 14),
        ("Line Items", 11),
        ("Total", 14),
        ("Currency", 8),
        ("Source File", 28),
        ("Notes", 32),
        ("Added", 18),
        ("Updated", 18),
    ]
    _style_header_row(ws1, columns_po)

    for r_idx, rec in enumerate(records, 2):
        col = 1
        def put(value, number_format=None):
            nonlocal col
            c = ws1.cell(row=r_idx, column=col, value=value)
            if number_format:
                c.number_format = number_format
            col += 1
            return c
        put(rec.get("po_number", ""))
        put(rec.get("po_date", ""))
        put(rec.get("revision", ""))
        put(rec.get("customer", ""))
        put(rec.get("supplier", ""))
        put(rec.get("supplier_code", ""))
        put(_flatten(rec.get("bill_to")))
        put(_flatten(rec.get("ship_to")))
        put(rec.get("payment_terms", ""))
        put(rec.get("freight_terms", ""))
        put(rec.get("ship_via", ""))
        put(rec.get("fob_terms", ""))
        put(rec.get("buyer", ""))
        put(rec.get("buyer_email", ""))
        put(rec.get("buyer_phone", ""))
        put(rec.get("receiving_contact", ""))
        put(rec.get("receiving_contact_phone", ""))
        put(rec.get("quote_number", ""))
        put(rec.get("contract_number", ""))
        put(len(rec.get("line_items") or []))
        put(float(rec.get("total") or 0), number_format='"$"#,##0.00')
        put(rec.get("currency", "USD"))
        put(rec.get("filename", ""))
        put(_flatten(rec.get("notes")))
        put(rec.get("added_at", ""))
        put(rec.get("updated_at", ""))

    # ---------- Sheet 2: Line Items ----------
    ws2 = wb.create_sheet("Line Items")

    columns_lines = [
        ("PO Number", 16),
        ("Customer", 24),
        ("Supplier", 24),
        ("Line", 6),
        ("Customer Part", 16),
        ("Vendor Part", 20),
        ("Description", 44),
        ("Quantity", 9),
        ("UOM", 6),
        ("Unit Price", 13),
        ("Amount", 14),
        ("Required Date", 14),
        ("Notes", 32),
    ]
    _style_header_row(ws2, columns_lines)

    r_idx = 2
    for rec in records:
        for it in rec.get("line_items") or []:
            ws2.cell(row=r_idx, column=1, value=rec.get("po_number", ""))
            ws2.cell(row=r_idx, column=2, value=rec.get("customer", ""))
            ws2.cell(row=r_idx, column=3, value=rec.get("supplier", ""))
            ws2.cell(row=r_idx, column=4, value=int(it.get("line") or 0))
            ws2.cell(row=r_idx, column=5, value=it.get("customer_part", ""))
            ws2.cell(row=r_idx, column=6, value=it.get("vendor_part", ""))
            ws2.cell(row=r_idx, column=7, value=it.get("description", ""))
            ws2.cell(row=r_idx, column=8, value=float(it.get("quantity") or 0))
            ws2.cell(row=r_idx, column=9, value=it.get("uom", ""))
            unit_cell = ws2.cell(row=r_idx, column=10, value=float(it.get("unit_price") or 0))
            unit_cell.number_format = '"$"#,##0.00'
            amt_cell = ws2.cell(row=r_idx, column=11, value=float(it.get("amount") or 0))
            amt_cell.number_format = '"$"#,##0.00'
            ws2.cell(row=r_idx, column=12, value=it.get("required_date", ""))
            ws2.cell(row=r_idx, column=13, value=_flatten(it.get("notes")))
            r_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
